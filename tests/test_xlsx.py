from openpyxl import load_workbook
import os.path


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    workbook = load_workbook(
        os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources/file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == 'Mara'